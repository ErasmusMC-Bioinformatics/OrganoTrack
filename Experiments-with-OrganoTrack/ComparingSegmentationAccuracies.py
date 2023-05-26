from pathlib import Path
from OrganoTrack.Detecting import SegmentWithOrganoSegPy
from OrganoTrack.Evaluating import EvaluateSegmentationAccuracy
import numpy as np
from OrganoTrack.Importing import ReadImages
from OrganoTrack.Displaying import ExportImageWithContours
import cv2 as cv
from datetime import datetime
import os
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import openpyxl
from scipy.stats import ttest_ind
import matplotlib.patches as mpatches
from statistics import stdev

'''
    With this file, boxplots can be generated, comparing the segmentation accuracies for different methods on different
    datasets. It is required, however, that each dataset directory is divided into "original", "groundTruth", and 
    "predicted".
'''


def SaveOverlays(gtAndPredictionOverlay, predictionOutlineOverlay, exportPath, imagePath):
    gtAndPredictionOverlayExportPath = exportPath / 'gtAndPredictionOverlay'
    predictionOutlineOverlayExportPath = exportPath / 'predictionOutlineOverlay'

    if not os.path.exists(gtAndPredictionOverlayExportPath):
        os.mkdir(gtAndPredictionOverlayExportPath)
    if not os.path.exists(predictionOutlineOverlayExportPath):
        os.mkdir(predictionOutlineOverlayExportPath)

    cv.imwrite(str(gtAndPredictionOverlayExportPath / imagePath.name), gtAndPredictionOverlay)
    cv.imwrite(str(predictionOutlineOverlayExportPath / imagePath.name), predictionOutlineOverlay)


# source: https://stackoverflow.com/questions/11517986/indicating-the-statistically-significant-difference-in-bar-graph
def barplot_annotate_brackets(num1, num2, data, center, height, yerr=None, dh=.05, barh=.05, fs=None, maxasterix=None):
    """
    Annotate barplot with p-values.

    :param num1: number of left bar to put bracket over
    :param num2: number of right bar to put bracket over
    :param data: string to write or number for generating asterixes
    :param center: centers of all bars (like plt.bar() input)
    :param height: heights of all bars (like plt.bar() input)
    :param yerr: yerrs of all bars (like plt.bar() input)
    :param dh: height offset over bar / bar + yerr in axes coordinates (0 to 1)
    :param barh: bar height in axes coordinates (0 to 1)
    :param fs: font size
    :param maxasterix: maximum number of asterixes to write (for very small p-values)
    """

    if type(data) is str:
        text = data
    else:
        # * is p < 0.05
        # ** is p < 0.005
        # *** is p < 0.0005
        # etc.
        text = ''
        p = .05

        while data < p:
            text += '*'
            p /= 10.

            if maxasterix and len(text) == maxasterix:
                break

        if len(text) == 0:
            text = 'n. s.'

    lx, ly = center[num1], height[num1]
    rx, ry = center[num2], height[num2]

    if yerr:
        ly += yerr[num1]
        ry += yerr[num2]

    ax_y0, ax_y1 = plt.gca().get_ylim()
    dh *= (ax_y1 - ax_y0)
    barh *= 0.2*(ax_y1 - ax_y0)

    y = max(ly, ry) + dh

    barx = [lx, lx, rx, rx]
    bary = [y, y+barh, y+barh, y]
    mid = ((lx+rx)/2, y+barh)

    plt.plot(barx, bary, c='black')

    kwargs = dict(ha='center', va='bottom')
    if fs is not None:
        kwargs['fontsize'] = fs

    plt.text(*mid, text, **kwargs)

def GetDatasetsDirs(datasets):
    datasetDirs = {'Bladder cancer': Path('/home/franz/Documents/mep/data/for-creating-OrganoTrack/training-dataset/preliminary-gt-dataset'),
                   'Mouse': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/MouseOrganoids'),
                   'Salivary ACC': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/ACCOrganoids'),
                   'Colon': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/ColonOrganoids'),
                   'Lung': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/LungOrganoids'),
                   'PDAC': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/PDACOrganoids')}

    datasetsDirs = dict()
    for dataset in datasets:
        datasetsDirs[dataset] = datasetDirs[dataset]
    return datasetsDirs

def LoadImages(datasetsDirs, predictionMethods, extraOrganoTrackBlur=False, blurSize=3):

    datasetsGroundTruthsAndPredictions = dict()

    for dataset in list(datasetsDirs.keys()):
        oneDatasetGroundTruthsAndPredictions = dict()
        datasetDir = datasetsDirs[dataset]

        groundTruthDir = datasetDir / 'groundTruth'
        groundTruthImages, groundTruthImagesNames = ReadImages(groundTruthDir)
        oneDatasetGroundTruthsAndPredictions['groundTruth'] = [groundTruthImages, groundTruthImagesNames]

        originalDir = datasetDir / 'original'
        originalImages, imagesNames = ReadImages(originalDir)
        oneDatasetGroundTruthsAndPredictions['original'] = [originalImages, imagesNames]

        for method in predictionMethods:
            predictionDir = datasetDir / 'predictions' / (method + '-segmented')
            if method == 'OrganoTrack' and not os.path.exists(predictionDir):
                segParams = [0.5, 250, 150, extraOrganoTrackBlur, blurSize]
                exportPath = datasetDir / 'predictions'
                saveSegParams = [True, exportPath, imagesNames]
                predictionImages = SegmentWithOrganoSegPy(originalImages, segParams, saveSegParams)
            else:
                predictionImages, imagesNames = ReadImages(predictionDir)

            oneDatasetGroundTruthsAndPredictions[method] = [predictionImages, imagesNames]

        datasetsGroundTruthsAndPredictions[dataset] = oneDatasetGroundTruthsAndPredictions

    return datasetsGroundTruthsAndPredictions

def ViewLoadedImages(datasetsGtAndPreds):
    datasets = list(datasetsGtAndPreds.keys())
    methods = list(datasetsGtAndPreds[datasets[0]].keys())
    for dataset in datasets:
        for method in methods:
            images = datasetsGtAndPreds[dataset][method][0]
            for i, image in enumerate(images):
                cv.imshow(f'{dataset}, {method},{i}', image)
    cv.waitKey(0)

def CreateDirectoryToStoreOverlays(datasetDirectory, predictionMethod):
    exportPath = datasetDirectory / 'predictions'
    overlayExportPath = exportPath / (predictionMethod + '-overlays')
    if not os.path.exists(overlayExportPath):
        os.mkdir(overlayExportPath)
    return overlayExportPath

def CalculatePredictionScores(datasetsGtAndPreds, datasetsDirectories, predictionMethods):

    datasetsSegmentationScoresWithDiffMethods = dict()

    for dataset in list(datasetsDirectories.keys()):
        print(f'Evaluating the {dataset} dataset.')
        oneDatasetSegmentationScoresWithDiffMethods = dict()
        datasetDirectory = datasetsDirectories[dataset]
        groundTruthImages = datasetsGtAndPreds[dataset]['groundTruth'][0]
        originalImages = datasetsGtAndPreds[dataset]['original'][0]

        for method in predictionMethods:
            predictedImages = datasetsGtAndPreds[dataset][method][0]
            predictedImagesNames = datasetsGtAndPreds[dataset][method][1]

            # Create directory to store overlays
            overlayExportPath = CreateDirectoryToStoreOverlays(datasetDirectory, method)

            segmentationScores = np.zeros((len(groundTruthImages), 3))

            # Evaluating prediction against ground truth
            for i, (prediction, groundTruth, original) in enumerate(zip(predictedImages, groundTruthImages, originalImages)):
                segmentationScores[i], gtAndPredOverlay = EvaluateSegmentationAccuracy(prediction, groundTruth)
                predictionOutlineOverlay = ExportImageWithContours(original, prediction)
                SaveOverlays(gtAndPredOverlay, predictionOutlineOverlay,  overlayExportPath, predictedImagesNames[i])

            oneDatasetSegmentationScoresWithDiffMethods[method] = segmentationScores
        oneDatasetSegmentationScoresWithDiffMethods['imageNames'] = predictedImagesNames

        datasetsSegmentationScoresWithDiffMethods[dataset] = oneDatasetSegmentationScoresWithDiffMethods

    return datasetsSegmentationScoresWithDiffMethods

def ExportPredictionScores(datasetsPredictionScores, analysisFileName, predictionMethods):
    # Exporting segmentation accuracies
    print(f'Exporting prediction scores to {analysisFileName}.')
    measureCount = 3
    datasets = list(datasetsPredictionScores.keys())

    with pd.ExcelWriter(str(analysisFileName.absolute())) as writer:

        for dataset in datasets:

            for j, method in enumerate(predictionMethods):
                methodSegmentationScores = datasetsPredictionScores[dataset][method]

                indexNames = []
                imagePaths = datasetsPredictionScores[dataset]['imageNames']
                for i in range(len(imagePaths)):
                    indexNames.append(imagePaths[i].name)

                df = pd.DataFrame(methodSegmentationScores, index=indexNames, columns=['f1', 'iou', 'dice'])
                df.to_excel(writer, sheet_name=dataset, startrow=1, startcol=j * (measureCount + 2))
                # Within .to_excel(), startrow/col are 0-indexed. Startcol calculated to fit df's next to each other

                writer.sheets[dataset].cell(row=1, column=j*(measureCount+2) + 1).value = str(method)
                # Within .cell(), row and column are 1-indexed

def LoadPredictionScoreAnalysis(analysisFilePath):
    xls = openpyxl.load_workbook(analysisFilePath)
    datasets = xls.sheetnames

    datasetsPredictionScores = dict()
    for dataset in datasets:
        oneDatasetMethodsPredictionsDf = pd.read_excel(analysisFilePath, sheet_name=str(dataset), header=None)

        oneDatasetMethodsPredictions = dict()

        firstRow = oneDatasetMethodsPredictionsDf.iloc[0]
        methods = firstRow.dropna().tolist()
        methodIndeces = list(firstRow.notna()[firstRow.notna() == True].index)

        for i, method in enumerate(methods):
            methodPredictions = np.array(oneDatasetMethodsPredictionsDf.iloc[2:, methodIndeces[i]+1:methodIndeces[i]+4], dtype=np.float64)
            oneDatasetMethodsPredictions[method] = methodPredictions

        datasetsPredictionScores[dataset] = oneDatasetMethodsPredictions

    return datasetsPredictionScores

def PlotPredictionAccuracies(datasetsPredictionScores, predictionMethods):
    datasets = list(datasetsPredictionScores.keys())
    # https: // matplotlib.org / stable / gallery / statistics / boxplot_demo.html
    measures = ['F1', 'IOU', 'Dice']
    measureIndex = [0, 1, 2]

    plt.rcParams.update({'font.size': 20})

    # Plotting colours
    methodColours = {'OrganoTrack': 'royalblue',
                     'OrganoID': 'darkorchid',
                     'Farhan1': 'darkgreen',
                     'Farhan2': 'seagreen'}


    for segAccuracyMeasure, index in zip(measures, measureIndex):
        plotTicks, boxplotPositionsPerMethod = ComputeBoxplotTicksAndPositions(predictionMethods, len(datasets))
        fig, ax = plt.subplots(figsize=(20, 6))
        legend_patches = []

        for method in predictionMethods:
            SegmentationAccuracyForOneMethodAcrossDatasets = []

            for dataset in datasets:
                SegmentationAccuracyForOneMethodAcrossDatasets.append(datasetsPredictionScores[dataset][method][:, index])

            ax.boxplot(SegmentationAccuracyForOneMethodAcrossDatasets, positions=boxplotPositionsPerMethod[method],
                       patch_artist=True, boxprops=dict(facecolor=methodColours[method]))

            legend_patch = mpatches.Patch(facecolor=methodColours[method], label=method)
            legend_patches.append(legend_patch)

        ax.set_ylabel(f'{segAccuracyMeasure} score')
        ax.set_xlabel('Organoid datasets')
        ax.set_ylim(0, 100)
        ax.set_xticks(plotTicks)
        ax.set_xticklabels(datasets)
        ax.legend(handles=legend_patches, bbox_to_anchor=(0.1, 0.5))
        plt.tight_layout()
        # valuesJitter = [np.random.normal(900, 10, 5), np.random.normal(1100, 10, 5)]
        # palette = [baselineColor, 'royalblue', 'r', 'c', 'm', 'k']
        # for jitter, val, c in zip(valuesJitter, data, palette):
        #     ax.scatter(jitter, val, alpha=0.4, color=c)
        fig.show()


def ComputeBoxplotTicksAndPositions(methods, numberOfDatasets):

    positionsOfBoxplotsForEachMethod = dict()
    numberOfMethods = len(methods)

    boxplotCharacteristics = {'distanceToFirstBoxFromYAxis': 2, 'boxWidth': 0.5, 'spaceBetweenBoxes': 0.5}

    boxplotXAxisTicks = ComputeBoxplotXAxisTicks(numberOfMethods, numberOfDatasets, boxplotCharacteristics)

    boxplotsPositionsForAllMethods = ComputeBoxplotPositionsForAllMethods(boxplotXAxisTicks, numberOfDatasets, methods,
                                                          boxplotCharacteristics)

    return boxplotXAxisTicks, boxplotsPositionsForAllMethods


def ComputeBoxplotXAxisTicks(numMethods, numDatasets, boxplotCharacteristics):
    distanceToFirstBoxFromYAxis = boxplotCharacteristics['distanceToFirstBoxFromYAxis']
    boxWidth = boxplotCharacteristics['boxWidth']
    spaceBetweenBoxes = boxplotCharacteristics['spaceBetweenBoxes']

    firstTick = distanceToFirstBoxFromYAxis + (numMethods/2)*boxWidth + ((numMethods-1)/2)*spaceBetweenBoxes
    distanceBetweenTicks = distanceToFirstBoxFromYAxis + numMethods*boxWidth + (numMethods-1)*spaceBetweenBoxes

    boxplotXAxisTicks = np.zeros(numDatasets)
    boxplotXAxisTicks[0] = firstTick
    for i in range(1, len(boxplotXAxisTicks)):
        boxplotXAxisTicks[i] = boxplotXAxisTicks[i-1] + distanceBetweenTicks

    return boxplotXAxisTicks

def TestComputePlotTicks():

    boxplotCharacteristics = {'distanceToFirstBoxFromYAxis': 20, 'boxWidth': 2, 'spaceBetweenBoxes': 0.5}
    numMethods = 3
    numDatasets = 4
    plotTicks = ComputeBoxplotXAxisTicks(numMethods, numDatasets, boxplotCharacteristics)

def ComputeBoxplotPositionsForAllMethods(plotTicks, numDatasets, methods, boxplotCharacteristics):
    boxWidth = boxplotCharacteristics['boxWidth']
    spaceBetweenBoxes = boxplotCharacteristics['spaceBetweenBoxes']

    numberOfMethods = len(methods)
    boxplotPositionsForAllMethods = dict()

    distanceToMethodBoxplotFromXAxisTick = np.zeros(numberOfMethods)
    distanceToMethodBoxplotFromXAxisTick[0] = -((numberOfMethods-1)/2)*(boxWidth+spaceBetweenBoxes)

    for i in range(1, len(distanceToMethodBoxplotFromXAxisTick)):
        distanceToMethodBoxplotFromXAxisTick[i] = distanceToMethodBoxplotFromXAxisTick[i-1]+(boxWidth+spaceBetweenBoxes)

    for i, method in enumerate(methods):
        boxplotPositionsForOneMethod = np.array([distanceToMethodBoxplotFromXAxisTick[i] + plotTicks[j]
                                                 for j in range(numDatasets)])
        boxplotPositionsForAllMethods[method] = boxplotPositionsForOneMethod

    return boxplotPositionsForAllMethods

def OrganoTrackVsHarmony():  # one dataset
    datasets = ['Bladder cancer']
    predictors = ['Harmony', 'OrganoTrack']
    analysisDir = Path('/home/franz/Documents/mep/report/results/segmentation-analysis')
    analysisFilePath = CreateAnalysisFileName(datasets, predictors, analysisDir)
    analysisExists = os.path.exists(analysisFilePath)
    toBlur = False

    if not analysisExists:
        datasetsDirs = GetDatasetsDirs(datasets)
        datasetsGtAndPreds = LoadImages(datasetsDirs, predictors, toBlur, 3)
        # ViewLoadedImages(datasetsGtAndPreds)
        datasetsPredictionScores = CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictors)
        ExportPredictionScores(datasetsPredictionScores, analysisFilePath, predictors)
    else:
        datasetsPredictionScores = LoadPredictionScoreAnalysis(analysisFilePath)

    # PlotPredictionAccuracies(datasetsPredictionScores, predictors)

    # Plotting
    x = np.array([1000])
    measures = ['F1', 'IOU', 'Dice']
    measureIndex = [0, 1, 2]
    boxWidth = 100
    plt.rcParams.update({'font.size': 20})
    baselineColor = 'indianred'
    measureSigPlot = {'F1': [87, 80], 'IOU': [80, 75], 'Dice': [87, 80]}
    for measure, index in zip(measures, measureIndex):
        data = []
        fig, ax = plt.subplots()
        for dataset in datasets:
            data1 = np.array(datasetsPredictionScores[dataset]['Harmony'][:, index]).T
            data.append(data1)
            ax.boxplot(data[0], positions=x-100, showfliers=False, widths=boxWidth) # one box plot corresponds to one method, not one dataset
            data2 = np.array(datasetsPredictionScores[dataset]['OrganoTrack'][:, index]).T
            data.append(data2)
            ax.boxplot(data[1], positions=x+100, showfliers=False, widths=boxWidth)


        ax.set_ylabel(f'{measure} score')
        ax.set_ylim(0, 100)
        labels = [item.get_text() for item in ax.get_xticklabels()]
        labels[0] = 'Baseline'  # change to OrganoTrack and baseline
        labels[1] = 'OrganoTrack'
        ax.set_xticklabels(labels)
        ax.set_xlim(800, 1200)
        valuesJitter = [np.random.normal(900, 10, 5), np.random.normal(1100, 10, 5)]
        palette = [baselineColor, 'royalblue', 'r', 'c', 'm', 'k']
        for jitter, val, c in zip(valuesJitter, data, palette):
            ax.scatter(jitter, val, alpha=0.4, color=c)
        plt.tight_layout()
        # barplot_annotate_brackets(0, 1, .25, [900, 1100], measureSigPlot[measure])
        # fig.show()


        measureSigPlot2 = {'F1': [87, 80], 'IOU': [80, 75], 'Dice': [87, 80]}
        fig1, ax1 = plt.subplots(figsize=(7, 6))
        ax1.bar(1, np.average(data[0]), yerr=stdev(data[0]), capsize=5, color=baselineColor)
        ax1.bar(2, np.average(data[1]), yerr=stdev(data[1]), capsize=5, color='royalblue')
        ax1.set_ylabel(f'{measure} score')
        ax1.set_ylim([0, 100])
        ax1.set_xticks([1, 2])
        ax1.set_xticklabels(['Baseline', 'OrganoTrack'])
        barplot_annotate_brackets(0, 1, .25, [1, 2], measureSigPlot2[measure])
        plt.tight_layout()
        fig1.show()

def OrganoTrackVsOrganoID():
    datasets = ['EMC-cisplatin', 'OrganoID-Mouse', 'OrganoID-Original']
    predictors = ['OrganoTrack', 'OrganoID']
    analysisDir = Path('/home/franz/Documents/mep/results/segmentation-analysis')

    analysisFile = analysisDir / (datasets[0] + '-' + predictors[0] + '-' + predictors[1] + '.xlsx')
    analysisExists = os.path.exists(analysisFile)

    # if not os.path.exists(analysisFile):

    if not analysisExists:
        datasetsDirs = GetDatasetsDirs(datasets)
        datasetsGtAndPreds = LoadImages(datasetsDirs, predictors, False, 3)
        # ViewLoadedImages(datasetsGtAndPreds)
        datasetsPredictionScores = CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictors)
        ExportPredictionScores(datasetsPredictionScores, analysisFile, predictors)
    else:
        datasetsPredictionScores = LoadPredictionScoreAnalysis(analysisFile)

    PlotPredictionAccuracies(datasetsPredictionScores, predictors)

    # # Plotting
    # x = np.array([1000, 2000, 3000])  # 3 datasets
    # measures = ['F1', 'IOU', 'Dice']
    # measureIndex = [0, 1, 2]
    # boxWidth = 100
    # plt.rcParams.update({'font.size': 15})
    #
    # # https://stackoverflow.com/questions/14952401/creating-double-boxplots-i-e-two-boxes-for-each-x-value
    # for measure, index in zip(measures, measureIndex):
    #     fig, ax = plt.subplots()
    #     data1 = []
    #     data2 = []
    #     for dataset in datasets:
    #         data1.append(datasetsPredictionScores[dataset]['OrganoTrack'][:, index])
    #         data2.append(datasetsPredictionScores[dataset]['OrganoID'][:, index])
    #     ax.boxplot(data1, positions=x-100, showfliers=False, widths=boxWidth) # one box plot corresponds to one method, not one dataset
    #     ax.boxplot(data2, positions=x+100,
    #                    showfliers=False, widths=boxWidth)   # one box plot corresponds to one method, not one dataset
    #         # fill with colors, https://matplotlib.org/stable/gallery/statistics/boxplot_color.html
    #         # colors = ['lightblue', 'lightgreen']
    #         # for bplot in bplot1:
    #         #     for patch, color in zip(bplot['boxes'], colors):
    #         #         patch.set_facecolor(color)
    #
    #     ax.set_ylabel(f'{measure} score')
    #     ax.set_ylim(0, 100)
    #     plt.xticks(x)
    #     labels = [item.get_text() for item in ax.get_xticklabels()]
    #     labels[0] = 'EMC\npreliminary\ndataset'
    #     labels[1] = 'OrganoID\nmouse organoids\ndataset'
    #     labels[2] = 'OrganoID\noriginal organoids\ndataset'
    #
    #     ax.set_xticklabels(labels)
    #     ax.set_xlim(800, 3200)
    #
    #     ax.set_title(f'{measure} score') #  for OrganoTrack and the baseline on a sample of the EMC dataset
    #     palette = ['b', 'g', 'r', 'c', 'm', 'k']
    #     # for i in range(len(conditionsNewOrder)):
    #     #     xs.append(np.random.normal(i + 1, 0.04, areaDFsSorted[i]['Norm Frac Growth'].values.shape[0]))
    #     #
    #     # for x, val, c in zip(xs, normFracGrowthValues, palette):
    #     #     ax.scatter(x, val, alpha=0.4, color=c)
    #     plt.tight_layout()
    #     fig.show()
    #     print('h')


def OrganoTrackBlurring():
    datasets = ['EMC-cisplatin']
    predictors = ['OrganoTrack']
    analysisDir = Path('/home/franz/Documents/mep/results/segmentation-analysis')
    analysisFile = analysisDir / (datasets[0] + '-' + predictors[0] + '-blurring.xlsx')
    analysisExists = os.path.exists(analysisFile)
    blurSizes = [3, 5, 7, 9, 11, 13, 15]
    datasetsDirs = GetDatasetsDirs(datasets)
    for blurSize in blurSizes:
        datasetsGtAndPreds = LoadImages(datasetsDirs, predictors, True, blurSize)
        # gets GT and predictions ^
        datasetsPredictionScores = CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictors)

    # ViewLoadedImages(datasetsGtAndPreds)

def CreateAnalysisFileName(datasetsNames, predictorsNames, analysisDir):

    modifieddatasetsNames = [name.replace(' ', '-') if ' ' in name else name for name in datasetsNames]
    datasetsCombinedString = '_'.join(modifieddatasetsNames) + '-datasets'
    predictorsCombinedString = '_'.join(predictorsNames) + '-methods'
    analysisFilePath = analysisDir / (datasetsCombinedString + '-' + predictorsCombinedString + '.xlsx')

    return analysisFilePath

def OrganoTrackVsOrganoIDvsFarhan():
    # To compare fairly with OrganoSeg, you'll need to remove all border objects from all other images.
    datasets = ['Bladder cancer', 'Mouse', 'Salivary ACC', 'Colon', 'Lung', 'PDAC']
    predictors = ['OrganoTrack', 'OrganoID', 'Farhan1', 'Farhan2']
    analysisDir = Path('/home/franz/Documents/mep/report/results/segmentation-analysis')

    analysisFilePath = CreateAnalysisFileName(datasets, predictors, analysisDir)

    analysisExists = os.path.exists(analysisFilePath)

    if not analysisExists:
        datasetsDirs = GetDatasetsDirs(datasets)
        datasetsGtAndPreds = LoadImages(datasetsDirs, predictors, extraOrganoTrackBlur=False, blurSize=3)
        datasetsPredictionScores = CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictors)
        ExportPredictionScores(datasetsPredictionScores, analysisFilePath, predictors)
    else:
        datasetsPredictionScores = LoadPredictionScoreAnalysis(analysisFilePath)

    PlotPredictionAccuracies(datasetsPredictionScores, predictors)


if __name__ == '__main__':
    # TestComputePlotTicks()
    OrganoTrackVsOrganoIDvsFarhan()
    # OrganoTrackVsHarmony()


