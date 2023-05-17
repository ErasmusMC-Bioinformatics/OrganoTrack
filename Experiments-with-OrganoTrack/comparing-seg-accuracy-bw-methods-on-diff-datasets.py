from pathlib import Path
from OrganoTrack.Detecting import SegmentWithOrganoSegPy
from OrganoTrack.Evaluating import EvaluateSegmentationAccuracy
import numpy as np
from OrganoTrack.Importing import ReadImages
from OrganoTrack.Displaying import DisplayImages
import cv2 as cv
from datetime import datetime
import os
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import openpyxl

'''
    With this file, boxplots can be generated, comparing the segmentation accuracies for different methods on different
    datasets. It is required, however, that each dataset directory is divided into "original", "groundTruth", and 
    "predicted".
'''


def SaveOverlay(overlay, exportPath, imagePath):
    cv.imwrite(str(exportPath / imagePath.name), overlay)


def GetDatasetsDirs(datasets):
    datasetDirs = {'EMC-preliminary': Path('/home/franz/Documents/mep/data/for-creating-OrganoTrack/training-dataset/preliminary-gt-dataset'),
                   'OrganoID-Mouse': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/MouseOrganoids'),
                   'OrganoID-Original': Path('/home/franz/Documents/mep/data/published-data/OrganoID-data/combinedForOrganoTrackTesting/OriginalData')}
    datasetsDirs = dict()
    for dataset in datasets:
        datasetsDirs[dataset] = datasetDirs[dataset]
    return datasetsDirs

def LoadImages(datasetsDirs, predictionMethods):

    datasetsGroundTruthsAndPredictions = dict()

    for dataset in list(datasetsDirs.keys()):
        oneDatasetGroundTruthsAndPredictions = dict()
        datasetDir = datasetsDirs[dataset]

        groundTruthDir = datasetDir / 'groundTruth'
        groundTruthImages, groundTruthImagesNames = ReadImages(groundTruthDir)
        oneDatasetGroundTruthsAndPredictions['groundTruth'] = [groundTruthImages, groundTruthImagesNames]

        for method in predictionMethods:
            predictionDir = datasetDir / 'predictions' / (method + '-segmented')
            if method == 'OrganoTrack' and not os.path.exists(predictionDir):
                originalImagesDir = datasetDir / 'original'
                originalImages, predictionImagesNames = ReadImages(originalImagesDir)
                segParams = [0.5, 250, 150]
                exportPath = datasetDir / 'predictions'
                saveSegParams = [True, exportPath, predictionImagesNames]
                predictionImages = SegmentWithOrganoSegPy(originalImages, segParams, saveSegParams)
            else:
                predictionImages, predictionImagesNames = ReadImages(predictionDir)

            oneDatasetGroundTruthsAndPredictions[method] = [predictionImages, predictionImagesNames]

        datasetsGroundTruthsAndPredictions[dataset] = oneDatasetGroundTruthsAndPredictions

    return datasetsGroundTruthsAndPredictions

def ViewLoadedImages(datasetsGtAndPreds):
    datasets = list(datasetsGtAndPreds.keys())
    methods = list(datasetsGtAndPreds[datasets[0]].keys())
    for dataset in datasets:
        for method in methods:
            images = datasetsGtAndPreds[dataset][method][0]
            for i, image in enumerate(images):
                cv.imshow(f'{dataset}, {method},{i}', image)
    cv.waitKey(0)


def CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictionMethods):

    datasetsSegScoresWithDiffMethods = dict()

    for dataset in list(datasetsDirs.keys()):
        segScoresWithDiffMethods = dict()
        datasetDir = datasetsDirs[dataset]
        groundTruthImages = datasetsGtAndPreds[dataset]['groundTruth'][0]

        for method in predictionMethods:
            predictedImages = datasetsGtAndPreds[dataset][method][0]
            predictedImagesNames = datasetsGtAndPreds[dataset][method][1]

            # Create directory to store overlays
            exportPath = datasetDir / 'predictions'
            overlayExportPath = exportPath / (method + '-overlay')
            if not os.path.exists(overlayExportPath):
                os.mkdir(overlayExportPath)

            segmentationScores = np.zeros((len(groundTruthImages), 3))

            # Evaluating prediction against ground truth
            for i, (prediction, groundTruth) in enumerate(zip(predictedImages, groundTruthImages)):
                segmentationScores[i], overlay = EvaluateSegmentationAccuracy(prediction, groundTruth)
                SaveOverlay(overlay, overlayExportPath, predictedImagesNames[i])

            segScoresWithDiffMethods[method] = segmentationScores
        segScoresWithDiffMethods['imageNames'] = predictedImagesNames

        datasetsSegScoresWithDiffMethods[dataset] = segScoresWithDiffMethods

    return datasetsSegScoresWithDiffMethods

def ExportPredictionScores(datasetsPredictionScores, analysisFileName, predictionMethods):
    # Exporting segmentation accuracies
    measureCount = 3
    datasets = list(datasetsPredictionScores.keys())

    with pd.ExcelWriter(str(analysisFileName.absolute())) as writer:

        for dataset in datasets:

            for j, method in enumerate(predictionMethods):
                methodSegmentationScores = datasetsPredictionScores[dataset][method]

                indexNames = []
                imagePaths = datasetsPredictionScores[dataset]['imageNames']
                for i in range(len(imagePaths)):
                    indexNames.append(imagePaths[i].name)

                df = pd.DataFrame(methodSegmentationScores, index=indexNames, columns=['f1', 'iou', 'dice'])
                df.to_excel(writer, sheet_name=dataset, startrow=1, startcol=j * (measureCount + 2))
                # Within .to_excel(), startrow/col are 0-indexed. Startcol calculated to fit df's next to each other

                writer.sheets[dataset].cell(row=1, column=j*(measureCount+2) + 1).value = str(method)
                # Within .cell(), row and column are 1-indexed

def LoadPredictionScoreAnalysis(analysisFilePath):
    xls = openpyxl.load_workbook(analysisFilePath)
    datasets = xls.sheetnames

    datasetsPredictionScores = dict()
    for dataset in datasets:
        oneDatasetMethodsPredictionsDf = pd.read_excel(analysisFilePath, sheet_name=str(dataset), header=None)

        oneDatasetMethodsPredictions = dict()

        firstRow = oneDatasetMethodsPredictionsDf.iloc[0]
        methods = firstRow.dropna().tolist()
        methodIndeces = list(firstRow.notna()[firstRow.notna() == True].index)

        for i, method in enumerate(methods):
            methodPredictions = oneDatasetMethodsPredictionsDf.iloc[2:, methodIndeces[i]+1:methodIndeces[i]+4].to_numpy()
            oneDatasetMethodsPredictions[method] = methodPredictions

        datasetsPredictionScores[dataset] = oneDatasetMethodsPredictions

    return datasetsPredictionScores



def PlotPredictionAccuracies(datasetsPredictionScores, predictionMethods):
    datasets = list(datasetsPredictionScores.keys())
    # https: // matplotlib.org / stable / gallery / statistics / boxplot_demo.html
    measures = ['F1', 'IOU', 'Dice']
    measureIndex = [0, 1, 2]

    plt.rcParams.update({'font.size': 15})

    # Plotting colours
    colours = list(mcolors.CSS4_COLORS.keys())

    for segAccuracyMeasure, index in zip(measures, measureIndex):
        fig, ax = plt.subplots()

        dataAcrossDatasetsAndMethods = []

        for method in predictionMethods:
            for dataset in datasets:
                dataAcrossDatasetsAndMethods.append(datasetsPredictionScores[dataset][method][:, index])

        # dataAcrossDatasetsAndMethods = np.array(dataAcrossDatasetsAndMethods).T
        ax.boxplot(dataAcrossDatasetsAndMethods)  # one box plot corresponds to one method, not one dataset

        ax.set_ylabel(f'{segAccuracyMeasure} score')
        ax.set_xlabel('Datasets')
        ax.set_ylim(0, 100)
        plt.tight_layout()
        fig.show()



def OrganoTrackVsHarmony():  # one dataset
    datasets = ['EMC-preliminary']
    predictors = ['Harmony', 'OrganoTrack']
    analysisDir = Path('/home/franz/Documents/mep/results/segmentation-analysis')
    analysisFile = analysisDir / (datasets[0] + '-' + predictors[0] + '-' + predictors[1] + '.xlsx')
    analysisExists = os.path.exists(analysisFile)

    if not analysisExists:
        datasetsDirs = GetDatasetsDirs(datasets)
        datasetsGtAndPreds = LoadImages(datasetsDirs, predictors)
        # ViewLoadedImages(datasetsGtAndPreds)
        datasetsPredictionScores = CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictors)
        ExportPredictionScores(datasetsPredictionScores, analysisFile, predictors)
    else:
        datasetsPredictionScores = LoadPredictionScoreAnalysis(analysisFile)

    PlotPredictionAccuracies(datasetsPredictionScores, predictors)

    # # Plotting
    # x = np.array([1000])
    # measures = ['F1', 'IOU', 'Dice']
    # measureIndex = [0, 1, 2]
    # boxWidth = 100
    # plt.rcParams.update({'font.size': 15})
    #
    #
    # for measure, index in zip(measures, measureIndex):
    #     fig, ax = plt.subplots()
    #     for dataset in datasets:
    #         data1 = np.array(datasetsPredictionScores[dataset]['Harmony'][:, index]).T
    #         ax.boxplot(data1, positions=x-100, showfliers=False, widths=boxWidth) # one box plot corresponds to one method, not one dataset
    #         data2 = np.array(datasetsPredictionScores[dataset]['OrganoTrack'][:, index]).T
    #         ax.boxplot(data2, positions=x+100,
    #                    showfliers=False, widths=boxWidth)   # one box plot corresponds to one method, not one dataset
    #         # fill with colors, https://matplotlib.org/stable/gallery/statistics/boxplot_color.html
    #         # colors = ['lightblue', 'lightgreen']
    #         # for bplot in bplot1:
    #         #     for patch, color in zip(bplot['boxes'], colors):
    #         #         patch.set_facecolor(color)
    #
    #     ax.set_ylabel(f'{measure} score')
    #     ax.set_ylim(0, 100)
    #     labels = [item.get_text() for item in ax.get_xticklabels()]
    #     labels[0] = 'Baseline'  # change to OrganoTrack and baseline
    #     labels[1] = 'OrganoTrack'
    #     ax.set_xticklabels(labels)
    #     ax.set_xlim(800, 1200)
    #
    #     # palette = ['b', 'g', 'r', 'c', 'm', 'k']
    #     # for x, val, c in zip(xs, normFracGrowthValues, palette):
    #     #     ax.scatter(x, val, alpha=0.4, color=c)
    #     plt.tight_layout()
    #     fig.show()

def OrganoTrackVsOrganoID():
    datasets = ['EMC-preliminary', 'OrganoID-Mouse', 'OrganoID-Original']
    predictors = ['OrganoTrack', 'OrganoID']
    analysisDir = Path('/home/franz/Documents/mep/results/segmentation-analysis')

    analysisFile = analysisDir / (datasets[0] + '-' + predictors[0] + '-' + predictors[1] + '.xlsx')
    analysisExists = os.path.exists(analysisFile)

    # if not os.path.exists(analysisFile):

    if not analysisExists:
        datasetsDirs = GetDatasetsDirs(datasets)
        datasetsGtAndPreds = LoadImages(datasetsDirs, predictors)
        # ViewLoadedImages(datasetsGtAndPreds)
        datasetsPredictionScores = CalculatePredictionScores(datasetsGtAndPreds, datasetsDirs, predictors)
        ExportPredictionScores(datasetsPredictionScores, analysisFile, predictors)
    else:
        datasetsPredictionScores = LoadPredictionScoreAnalysis(analysisFile)

    PlotPredictionAccuracies(datasetsPredictionScores, predictors)

    # # Plotting
    # x = np.array([1000, 2000, 3000])  # 3 datasets
    # measures = ['F1', 'IOU', 'Dice']
    # measureIndex = [0, 1, 2]
    # boxWidth = 100
    # plt.rcParams.update({'font.size': 15})
    #
    # # https://stackoverflow.com/questions/14952401/creating-double-boxplots-i-e-two-boxes-for-each-x-value
    # for measure, index in zip(measures, measureIndex):
    #     fig, ax = plt.subplots()
    #     data1 = []
    #     data2 = []
    #     for dataset in datasets:
    #         data1.append(datasetsPredictionScores[dataset]['OrganoTrack'][:, index])
    #         data2.append(datasetsPredictionScores[dataset]['OrganoID'][:, index])
    #     ax.boxplot(data1, positions=x-100, showfliers=False, widths=boxWidth) # one box plot corresponds to one method, not one dataset
    #     ax.boxplot(data2, positions=x+100,
    #                    showfliers=False, widths=boxWidth)   # one box plot corresponds to one method, not one dataset
    #         # fill with colors, https://matplotlib.org/stable/gallery/statistics/boxplot_color.html
    #         # colors = ['lightblue', 'lightgreen']
    #         # for bplot in bplot1:
    #         #     for patch, color in zip(bplot['boxes'], colors):
    #         #         patch.set_facecolor(color)
    #
    #     ax.set_ylabel(f'{measure} score')
    #     ax.set_ylim(0, 100)
    #     plt.xticks(x)
    #     labels = [item.get_text() for item in ax.get_xticklabels()]
    #     labels[0] = 'EMC\npreliminary\ndataset'
    #     labels[1] = 'OrganoID\nmouse organoids\ndataset'
    #     labels[2] = 'OrganoID\noriginal organoids\ndataset'
    #
    #     ax.set_xticklabels(labels)
    #     ax.set_xlim(800, 3200)
    #
    #     ax.set_title(f'{measure} score') #  for OrganoTrack and the baseline on a sample of the EMC dataset
    #     palette = ['b', 'g', 'r', 'c', 'm', 'k']
    #     # for i in range(len(conditionsNewOrder)):
    #     #     xs.append(np.random.normal(i + 1, 0.04, areaDFsSorted[i]['Norm Frac Growth'].values.shape[0]))
    #     #
    #     # for x, val, c in zip(xs, normFracGrowthValues, palette):
    #     #     ax.scatter(x, val, alpha=0.4, color=c)
    #     plt.tight_layout()
    #     fig.show()
    #     print('h')


if __name__ == '__main__':
    OrganoTrackVsOrganoID()


